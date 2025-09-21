"""
Purchase Order Excel Generator
生成與HTML相同格式的Excel檔案
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import io
import os
from PIL import Image as PILImage


class POExcelGenerator:
    """Purchase Order Excel Generator"""
    
    def generate_excel(self, purchase_order) -> bytes:
        """Generate Excel for purchase order"""
        wb = Workbook()
        ws = wb.active
        ws.title = "採購單"
        
        # Set column widths
        column_widths = [5, 20, 15, 12, 8, 8, 12, 12, 8]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Styles
        title_font = Font(name='Microsoft YaHei', size=20, bold=True)
        subtitle_font = Font(name='Arial', size=14, color='666666')
        header_font = Font(name='Microsoft YaHei', size=11, bold=True)
        normal_font = Font(name='Microsoft YaHei', size=10)
        total_font = Font(name='Microsoft YaHei', size=11, bold=True, color='0066CC')
        
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        gray_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
        light_gray_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
        blue_fill = PatternFill(start_color='E8F4F8', end_color='E8F4F8', fill_type='solid')
        
        thin_border = Border(
            left=Side(style='thin', color='999999'),
            right=Side(style='thin', color='999999'),
            top=Side(style='thin', color='999999'),
            bottom=Side(style='thin', color='999999')
        )
        
        current_row = 1
        
        # Try to add logo
        logo_paths = [
            os.path.join(os.path.dirname(__file__), '..', '..', '..', 'docs', 'TSIC_LOGO.png'),
            'docs/TSIC_LOGO.png'
        ]
        
        logo_added = False
        for logo_path in logo_paths:
            if os.path.exists(logo_path):
                try:
                    # Resize logo to fit
                    pil_img = PILImage.open(logo_path)
                    # Scale to 270px width (Excel uses 96 DPI)
                    width_px = 270
                    height_px = int(pil_img.height * (width_px / pil_img.width))
                    pil_img = pil_img.resize((width_px, height_px), PILImage.Resampling.LANCZOS)
                    
                    # Save to memory
                    img_buffer = io.BytesIO()
                    pil_img.save(img_buffer, format='PNG')
                    img_buffer.seek(0)
                    
                    # Add to Excel
                    img = Image(img_buffer)
                    img.width = width_px
                    img.height = height_px
                    ws.add_image(img, 'A1')
                    logo_added = True
                    current_row = 5  # Skip rows for logo
                    break
                except Exception as e:
                    print(f"Error adding logo: {e}")
        
        if not logo_added:
            # Add text logo
            ws.merge_cells(f'A1:C3')
            ws['A1'] = 'TSIC'
            ws['A1'].font = Font(name='Arial', size=48, bold=True, color='0066CC')
            ws['A1'].alignment = center_align
            current_row = 4
        
        # Title
        ws.merge_cells(f'D{current_row}:F{current_row}')
        ws[f'D{current_row}'] = '採購單'
        ws[f'D{current_row}'].font = title_font
        ws[f'D{current_row}'].alignment = center_align
        
        current_row += 1
        ws.merge_cells(f'D{current_row}:F{current_row}')
        ws[f'D{current_row}'] = 'PURCHASE ORDER'
        ws[f'D{current_row}'].font = subtitle_font
        ws[f'D{current_row}'].alignment = center_align
        
        current_row += 2
        
        # Info section
        info_data = [
            ['廠商名稱', purchase_order.supplier_name or '', '採購單號', purchase_order.purchase_order_no],
            ['廠商編號', purchase_order.supplier_id or '', '訂購日期', 
             purchase_order.order_date.strftime('%Y/%m/%d') if purchase_order.order_date else ''],
            ['廠商地址', purchase_order.supplier_address or '-', '報價單號', purchase_order.quotation_no or '-'],
            ['連絡電話', purchase_order.contact_phone or '-', '聯絡人', purchase_order.contact_person or '-']
        ]
        
        for row_data in info_data:
            ws[f'A{current_row}'] = row_data[0]
            ws[f'A{current_row}'].font = header_font
            ws[f'A{current_row}'].fill = light_gray_fill
            ws[f'A{current_row}'].border = thin_border
            ws.merge_cells(f'A{current_row}:B{current_row}')
            
            ws[f'C{current_row}'] = row_data[1]
            ws[f'C{current_row}'].font = normal_font
            ws[f'C{current_row}'].border = thin_border
            ws.merge_cells(f'C{current_row}:D{current_row}')
            
            ws[f'E{current_row}'] = row_data[2]
            ws[f'E{current_row}'].font = header_font
            ws[f'E{current_row}'].fill = light_gray_fill
            ws[f'E{current_row}'].border = thin_border
            ws.merge_cells(f'E{current_row}:F{current_row}')
            
            ws[f'G{current_row}'] = row_data[3]
            ws[f'G{current_row}'].font = normal_font
            ws[f'G{current_row}'].border = thin_border
            ws.merge_cells(f'G{current_row}:I{current_row}')
            
            current_row += 1
        
        current_row += 1
        
        # Items header
        headers = ['項次', '品名', '規格', '型號', '數量', '單位', '單價', '小計', '備註']
        for i, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=i, value=header)
            cell.font = header_font
            cell.fill = gray_fill
            cell.border = thin_border
            cell.alignment = center_align
        
        current_row += 1
        
        # Items
        items = purchase_order.items.all()
        for idx, item in enumerate(items, 1):
            ws.cell(row=current_row, column=1, value=idx).border = thin_border
            ws.cell(row=current_row, column=1).alignment = center_align
            
            ws.cell(row=current_row, column=2, value=item.item_name or '').border = thin_border
            ws.cell(row=current_row, column=3, value=item.item_specification or '-').border = thin_border
            ws.cell(row=current_row, column=4, value=item.item_model or '-').border = thin_border
            
            ws.cell(row=current_row, column=5, value=float(item.item_quantity)).border = thin_border
            ws.cell(row=current_row, column=5).alignment = right_align
            ws.cell(row=current_row, column=5).number_format = '#,##0'
            
            ws.cell(row=current_row, column=6, value=item.item_unit or '').border = thin_border
            ws.cell(row=current_row, column=6).alignment = center_align
            
            ws.cell(row=current_row, column=7, value=float(item.unit_price)).border = thin_border
            ws.cell(row=current_row, column=7).alignment = right_align
            ws.cell(row=current_row, column=7).number_format = '$#,##0'
            
            ws.cell(row=current_row, column=8, value=item.line_subtotal_int).border = thin_border
            ws.cell(row=current_row, column=8).alignment = right_align
            ws.cell(row=current_row, column=8).number_format = '$#,##0'
            
            ws.cell(row=current_row, column=9, value='').border = thin_border
            
            current_row += 1
        
        # Add one empty row if only one item
        if len(items) == 1:
            for i in range(1, 10):
                ws.cell(row=current_row, column=i, value='').border = thin_border
            current_row += 1
        
        current_row += 1
        
        # Summary section - matching HTML layout
        summary_start_row = current_row
        
        # Terms and Conditions title row
        ws.merge_cells(f'A{current_row}:E{current_row}')
        ws[f'A{current_row}'] = '注意事項 Terms and Conditions'
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].fill = gray_fill
        ws[f'A{current_row}'].border = thin_border
        ws[f'A{current_row}'].alignment = left_align
        
        # Subtotal on same row
        ws[f'F{current_row}'] = '小計 Subtotal'
        ws[f'F{current_row}'].font = header_font
        ws[f'F{current_row}'].fill = light_gray_fill
        ws[f'F{current_row}'].border = thin_border
        ws[f'F{current_row}'].alignment = right_align
        ws.merge_cells(f'F{current_row}:G{current_row}')
        
        ws[f'H{current_row}'] = f'NT$ {purchase_order.subtotal_int:,}'
        ws[f'H{current_row}'].font = normal_font
        ws[f'H{current_row}'].border = thin_border
        ws[f'H{current_row}'].alignment = right_align
        ws.merge_cells(f'H{current_row}:I{current_row}')
        
        current_row += 1
        
        # Terms content
        terms_content = [
            '1. 付款條件：月結 30 天',
            '2. 交貨期限：訂單確認後 14 個工作天',
            '3. 品質要求：須符合國家標準規範',
            '4. 驗收標準：貨到 7 日內完成驗收',
            '5. 保固期限：自驗收合格日起算一年'
        ]
        
        # Terms content rows
        terms_start_row = current_row
        for i, term in enumerate(terms_content):
            ws[f'A{current_row + i}'] = term
            ws[f'A{current_row + i}'].font = normal_font
            ws[f'A{current_row + i}'].alignment = left_align
            ws.merge_cells(f'A{current_row + i}:E{current_row + i}')
            
            # Apply border to merged cells
            for col in range(1, 6):
                ws.cell(row=current_row + i, column=col).border = thin_border
        
        # Tax row (aligned with first terms row)
        ws[f'F{terms_start_row}'] = '稅額 Tax (5%)'
        ws[f'F{terms_start_row}'].font = header_font
        ws[f'F{terms_start_row}'].fill = light_gray_fill
        ws[f'F{terms_start_row}'].border = thin_border
        ws[f'F{terms_start_row}'].alignment = right_align
        ws.merge_cells(f'F{terms_start_row}:G{terms_start_row}')
        
        ws[f'H{terms_start_row}'] = f'NT$ {float(purchase_order.tax_decimal1):,.1f}'
        ws[f'H{terms_start_row}'].font = normal_font
        ws[f'H{terms_start_row}'].border = thin_border
        ws[f'H{terms_start_row}'].alignment = right_align
        ws.merge_cells(f'H{terms_start_row}:I{terms_start_row}')
        
        # Total row (spans remaining height)
        total_start_row = terms_start_row + 1
        total_end_row = terms_start_row + len(terms_content) - 1
        
        ws.merge_cells(f'F{total_start_row}:G{total_end_row}')
        ws[f'F{total_start_row}'] = '總計 Total'
        ws[f'F{total_start_row}'].font = total_font
        ws[f'F{total_start_row}'].fill = blue_fill
        ws[f'F{total_start_row}'].alignment = Alignment(horizontal='right', vertical='center')
        
        ws.merge_cells(f'H{total_start_row}:I{total_end_row}')
        ws[f'H{total_start_row}'] = f'NT$ {purchase_order.grand_total_int:,}'
        ws[f'H{total_start_row}'].font = total_font
        ws[f'H{total_start_row}'].fill = blue_fill
        ws[f'H{total_start_row}'].alignment = Alignment(horizontal='right', vertical='center')
        
        # Apply borders to total cells
        for row in range(total_start_row, total_end_row + 1):
            for col in range(6, 10):
                ws.cell(row=row, column=col).border = thin_border
                ws.cell(row=row, column=col).fill = blue_fill
        
        current_row = terms_start_row + len(terms_content) + 2
        
        # Signature section - all on same row as HTML
        ws[f'A{current_row}'] = '製表人：'
        ws[f'A{current_row}'].font = header_font
        
        ws[f'B{current_row}'] = '_________________'
        ws.merge_cells(f'B{current_row}:C{current_row}')
        
        ws[f'D{current_row}'] = '日期：'
        ws[f'D{current_row}'].font = header_font
        
        ws[f'E{current_row}'] = '_________________'
        
        ws[f'F{current_row}'] = '採購主管：'
        ws[f'F{current_row}'].font = header_font
        
        ws[f'G{current_row}'] = '_________________'
        ws.merge_cells(f'G{current_row}:H{current_row}')
        
        ws[f'I{current_row}'] = '日期：_________________'
        ws[f'I{current_row}'].font = header_font
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer.read()