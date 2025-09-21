"""
Purchase Order File Generation Service
Handles Excel and PDF generation for purchase orders
"""
import os
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, Optional
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from decimal import Decimal
import tempfile
from app.models import PurchaseOrder, PurchaseOrderItem


class POGenerator:
    """Purchase Order File Generator"""
    
    def __init__(self):
        self.template_dir = Path(__file__).parent.parent / 'templates'
        self.excel_template = self.template_dir / 'po_template.xlsx'
        
        # Register Chinese font for PDF if available
        try:
            font_path = "C:/Windows/Fonts/msjh.ttc"  # Microsoft JhengHei
            if os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont('Chinese', font_path))
                self.chinese_font = 'Chinese'
            else:
                self.chinese_font = 'Helvetica'
        except:
            self.chinese_font = 'Helvetica'
    
    def generate_excel(self, purchase_order: PurchaseOrder) -> bytes:
        """Generate Excel file for purchase order"""
        # Load template or create new workbook
        if self.excel_template.exists():
            wb = openpyxl.load_workbook(self.excel_template)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            self._create_excel_template(ws)
        
        # Fill in purchase order data
        self._fill_excel_data(ws, purchase_order)
        
        # Save to bytes using BytesIO
        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        excel_data = buffer.getvalue()
        buffer.close()
        
        return excel_data
    
    def _create_excel_template(self, ws):
        """Create basic Excel template structure"""
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 20
        
        # Header
        ws['A1'] = '採購單 Purchase Order'
        ws['A1'].font = Font(size=18, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:H1')
        
        # Company info
        ws['A3'] = '公司名稱：'
        ws['B3'] = 'ERP 系統公司'
        ws['A4'] = '公司地址：'
        ws['B4'] = '台北市信義區信義路五段7號'
        ws['A5'] = '聯絡電話：'
        ws['B5'] = '02-8101-2345'
        
        # Purchase order info
        ws['F3'] = '採購單號：'
        ws['F4'] = '採購日期：'
        ws['F5'] = '交貨地址：'
        
        # Supplier info
        ws['A7'] = '供應商資訊 Supplier Information'
        ws['A7'].font = Font(bold=True)
        ws.merge_cells('A7:H7')
        
        ws['A8'] = '供應商名稱：'
        ws['A9'] = '供應商地址：'
        ws['A10'] = '聯絡人：'
        ws['E10'] = '聯絡電話：'
        
        # Items header
        row = 12
        headers = ['項次', '品名', '規格', '數量', '單位', '單價', '小計', '備註']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    
    def _fill_excel_data(self, ws, po: PurchaseOrder):
        """Fill Excel template with purchase order data"""
        # Purchase order info
        ws['G3'] = po.purchase_order_no
        ws['G4'] = po.order_date.strftime('%Y-%m-%d') if po.order_date else ''
        ws['G5'] = po.delivery_address or ''
        
        # Supplier info
        ws['B8'] = po.supplier_name
        ws['B9'] = po.supplier_address or ''
        ws['B10'] = po.contact_person or ''
        ws['F10'] = po.contact_phone or ''
        
        # Items
        items = po.items.all()
        start_row = 13
        for idx, item in enumerate(items, 1):
            row = start_row + idx - 1
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=item.item_name)
            ws.cell(row=row, column=3, value=item.item_specification or '')
            ws.cell(row=row, column=4, value=float(item.item_quantity))
            ws.cell(row=row, column=5, value=item.item_unit)
            ws.cell(row=row, column=6, value=float(item.unit_price))
            ws.cell(row=row, column=7, value=item.line_subtotal_int)
            ws.cell(row=row, column=8, value='')
            
            # Add borders
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Totals
        total_row = start_row + len(items) + 1
        ws.cell(row=total_row, column=6, value='小計：')
        ws.cell(row=total_row, column=6).alignment = Alignment(horizontal='right')
        ws.cell(row=total_row, column=7, value=po.subtotal_int)
        
        ws.cell(row=total_row+1, column=6, value='稅額 (5%)：')
        ws.cell(row=total_row+1, column=6).alignment = Alignment(horizontal='right')
        ws.cell(row=total_row+1, column=7, value=float(po.tax_decimal1))
        
        ws.cell(row=total_row+2, column=6, value='總計：')
        ws.cell(row=total_row+2, column=6).font = Font(bold=True)
        ws.cell(row=total_row+2, column=6).alignment = Alignment(horizontal='right')
        ws.cell(row=total_row+2, column=7, value=po.grand_total_int)
        ws.cell(row=total_row+2, column=7).font = Font(bold=True)
        
        # Notes
        notes_row = total_row + 4
        ws.cell(row=notes_row, column=1, value='備註：')
        ws.merge_cells(f'B{notes_row}:H{notes_row}')
        ws.cell(row=notes_row, column=2, value=po.notes or '')
        
        # Signature area
        sig_row = notes_row + 3
        ws.cell(row=sig_row, column=1, value='採購人員：')
        ws.cell(row=sig_row, column=3, value='_________________')
        ws.cell(row=sig_row, column=5, value='核准主管：')
        ws.cell(row=sig_row, column=7, value='_________________')
    
    def generate_pdf(self, purchase_order: PurchaseOrder) -> bytes:
        """Generate PDF file for purchase order"""
        # Create PDF in memory
        buffer = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        buffer_name = buffer.name
        buffer.close()  # Close the file handle before using it with SimpleDocTemplate
        
        doc = SimpleDocTemplate(buffer_name, pagesize=A4,
                                rightMargin=20*mm, leftMargin=20*mm,
                                topMargin=20*mm, bottomMargin=20*mm)
        
        # Container for PDF elements
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#333333'),
            alignment=TA_CENTER,
            fontName=self.chinese_font
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#333333'),
            fontName=self.chinese_font,
            spaceAfter=6
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=10,
            fontName=self.chinese_font
        )
        
        # Title
        elements.append(Paragraph('採購單 Purchase Order', title_style))
        elements.append(Spacer(1, 20))
        
        # Company and PO info table
        company_data = [
            ['公司名稱：', 'ERP 系統公司', '', '採購單號：', purchase_order.purchase_order_no],
            ['公司地址：', '台北市信義區信義路五段7號', '', '採購日期：', 
             purchase_order.order_date.strftime('%Y-%m-%d') if purchase_order.order_date else ''],
            ['聯絡電話：', '02-8101-2345', '', '交貨地址：', purchase_order.delivery_address or '']
        ]
        
        company_table = Table(company_data, colWidths=[60, 120, 40, 60, 120])
        company_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), self.chinese_font),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(company_table)
        elements.append(Spacer(1, 20))
        
        # Supplier info
        elements.append(Paragraph('供應商資訊 Supplier Information', heading_style))
        
        supplier_data = [
            ['供應商名稱：', purchase_order.supplier_name, '', ''],
            ['供應商地址：', purchase_order.supplier_address or '', '', ''],
            ['聯絡人：', purchase_order.contact_person or '', '聯絡電話：', purchase_order.contact_phone or ''],
            ['統一編號：', purchase_order.supplier_tax_id or '', '', '']
        ]
        
        supplier_table = Table(supplier_data, colWidths=[70, 150, 70, 110])
        supplier_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), self.chinese_font),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BOX', (0, 0), (-1, -1), 1, colors.grey),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BACKGROUND', (0, 0), (0, -1), colors.Color(0.9, 0.9, 0.9)),
            ('BACKGROUND', (2, 0), (2, -1), colors.Color(0.9, 0.9, 0.9)),
        ]))
        elements.append(supplier_table)
        elements.append(Spacer(1, 20))
        
        # Items
        elements.append(Paragraph('採購明細 Purchase Items', heading_style))
        
        # Items table
        items_data = [['項次', '品名', '規格', '數量', '單位', '單價', '小計']]
        
        items = purchase_order.items.all()
        for idx, item in enumerate(items, 1):
            items_data.append([
                str(idx),
                item.item_name,
                item.item_specification or '',
                str(float(item.item_quantity)),
                item.item_unit,
                f'{float(item.unit_price):,.2f}',
                f'{item.line_subtotal_int:,}'
            ])
        
        # Add empty rows if needed
        while len(items_data) < 6:
            items_data.append(['', '', '', '', '', '', ''])
        
        items_table = Table(items_data, colWidths=[30, 120, 80, 40, 40, 50, 50])
        items_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), self.chinese_font),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),
            ('ALIGN', (3, 1), (6, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.8, 0.8, 0.8)),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(items_table)
        elements.append(Spacer(1, 15))
        
        # Totals
        totals_data = [
            ['', '', '', '', '', '小計：', f'{purchase_order.subtotal_int:,}'],
            ['', '', '', '', '', '稅額 (5%)：', f'{float(purchase_order.tax_decimal1):,.1f}'],
            ['', '', '', '', '', '總計：', f'{purchase_order.grand_total_int:,}']
        ]
        
        totals_table = Table(totals_data, colWidths=[30, 120, 80, 40, 40, 50, 50])
        totals_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), self.chinese_font),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (5, 0), (5, -1), 'RIGHT'),
            ('ALIGN', (6, 0), (6, -1), 'RIGHT'),
            ('FONTSIZE', (5, 2), (6, 2), 12),
            ('TEXTCOLOR', (5, 2), (6, 2), colors.HexColor('#000000')),
            ('BACKGROUND', (5, 2), (6, 2), colors.Color(0.9, 0.9, 0.9)),
            ('BOX', (5, 0), (6, 2), 1, colors.grey),
            ('LINEABOVE', (5, 2), (6, 2), 2, colors.black),
        ]))
        elements.append(totals_table)
        elements.append(Spacer(1, 20))
        
        # Notes
        if purchase_order.notes:
            elements.append(Paragraph('備註 Notes', heading_style))
            elements.append(Paragraph(purchase_order.notes, normal_style))
            elements.append(Spacer(1, 20))
        
        # Terms and conditions
        terms = """
        付款條件：月結 30 天
        交貨期限：訂單確認後 14 個工作天
        品質要求：符合國家標準規範
        """
        elements.append(Paragraph('條款與條件 Terms and Conditions', heading_style))
        elements.append(Paragraph(terms, normal_style))
        elements.append(Spacer(1, 30))
        
        # Signature area
        sig_data = [
            ['採購人員：', '_________________', '', '核准主管：', '_________________'],
            ['', '', '', '', ''],
            ['日期：', '_________________', '', '日期：', '_________________']
        ]
        
        sig_table = Table(sig_data, colWidths=[60, 100, 40, 60, 100])
        sig_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), self.chinese_font),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ]))
        elements.append(sig_table)
        
        # Build PDF
        doc.build(elements)
        
        # Read PDF data
        with open(buffer_name, 'rb') as f:
            pdf_data = f.read()
        os.unlink(buffer_name)
        
        return pdf_data
    
    def get_preview_data(self, purchase_order: PurchaseOrder) -> Dict[str, Any]:
        """Get purchase order data for preview"""
        items = []
        for idx, item in enumerate(purchase_order.items.all(), 1):
            items.append({
                'index': idx,
                'item_name': item.item_name,
                'item_specification': item.item_specification or '',
                'item_quantity': float(item.item_quantity),
                'item_unit': item.item_unit,
                'unit_price': float(item.unit_price),
                'line_subtotal': item.line_subtotal_int,
                'source_request_order_no': item.source_request_order_no
            })
        
        return {
            'purchase_order_no': purchase_order.purchase_order_no,
            'supplier_name': purchase_order.supplier_name,
            'supplier_address': purchase_order.supplier_address,
            'contact_person': purchase_order.contact_person,
            'contact_phone': purchase_order.contact_phone,
            'supplier_tax_id': purchase_order.supplier_tax_id,
            'order_date': purchase_order.order_date.isoformat() if purchase_order.order_date else None,
            'quotation_no': purchase_order.quotation_no,
            'delivery_address': purchase_order.delivery_address,
            'creation_date': purchase_order.creation_date.isoformat() if purchase_order.creation_date else None,
            'notes': purchase_order.notes,
            'purchase_status': purchase_order.purchase_status,
            'subtotal_int': purchase_order.subtotal_int,
            'tax_decimal1': float(purchase_order.tax_decimal1),
            'grand_total_int': purchase_order.grand_total_int,
            'items': items,
            'company': {
                'name': 'ERP 系統公司',
                'address': '台北市信義區信義路五段7號',
                'phone': '02-8101-2345',
                'tax_id': '12345678'
            },
            'terms': {
                'payment': '月結 30 天',
                'delivery': '訂單確認後 14 個工作天',
                'quality': '符合國家標準規範'
            }
        }